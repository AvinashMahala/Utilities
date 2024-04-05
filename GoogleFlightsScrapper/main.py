from google_flight_analysis.scrape import Scrape, ScrapeObjects
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class FlightData:
    def __init__(self, direction, data):
        self.direction = direction
        self.departure_datetime = data.iloc[0]
        self.arrival_datetime = data.iloc[1]
        self.origin = data.iloc[2]
        self.destination = data.iloc[3]
        self.airlines = data.iloc[4]
        self.travel_time = data.iloc[5]
        self.price = data.iloc[6]
        self.num_stops = data.iloc[7]
        self.layover = data.iloc[8]
        self.access_date = data.iloc[9]
        self.co2_emission = data.iloc[10]
        self.emission_diff = data.iloc[11]

    def print_flight_data_obj(self):
        print("Direction:", self.direction)
        print("Departure Datetime:", self.departure_datetime)
        print("Arrival Datetime:", self.arrival_datetime)
        print("Origin:", self.origin)
        print("Destination:", self.destination)
        print("Airline(s):", self.airlines)
        print("Travel Time:", self.travel_time)
        print("Price ($):", self.price)
        print("Number of Stops:", self.num_stops)
        print("Layover:", self.layover)
        print("Access Date:", self.access_date)
        print("CO2 Emission (kg):", self.co2_emission)
        print("Emission Diff (%):", self.emission_diff)

    def to_list(self):
        return [self.direction, self.departure_datetime, self.arrival_datetime, self.origin, self.destination,
                self.airlines, self.travel_time, self.price, self.num_stops, self.layover,
                self.access_date, self.co2_emission, self.emission_diff]
def generate_filename(start_date, end_date, start_interval, end_interval, prefix='lowest_prices', extension='xlsx'):
    start_date_str = start_date.strftime('%Y%m%d')
    end_date_str = end_date.strftime('%Y%m%d')
    interval_str = f"{start_interval}_{end_interval}"
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{prefix}_{start_date_str}_{end_date_str}_{interval_str}_{timestamp_str}.{extension}"
    return filename
def write_to_excel(lowest_price_up, lowest_price_down, filename):
    try:
        # Load existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # Create a new Workbook if the file does not exist
        wb = Workbook()
        ws = wb.active
        # Write column headers
        headers = ['Direction', 'Departure datetime', 'Arrival datetime', 'Origin', 'Destination',
                   'Airline(s)', 'Travel Time', 'Price ($)', 'Num Stops', 'Layover',
                   'Access Date', 'CO2 Emission (kg)', 'Emission Diff (%)']
        ws.append(headers)

    # Write data for lowest price up
    ws.append(lowest_price_up.to_list())

    # Write data for lowest price down
    ws.append(lowest_price_down.to_list())

    # Save the workbook
    wb.save(filename)


def print_flight_data(flight_data):
    # Set option to display all columns
    pd.set_option('display.max_columns', None)
    # Define the new column names and order
    new_column_names = {
        'Departure datetime': 'Departure Date',
        'Arrival datetime': 'Arrival Date',
        'Origin': 'From',
        'Destination': 'To',
        'Airline(s)': 'Airline',
        'Travel Time': 'Duration',
        'Price ($)': 'Price',
        'Num Stops': 'Number of Stops',
        'Layover': 'Layover Time',
        'Access Date': 'Accessed Date',
        'CO2 Emission (kg)': 'CO2 Emission',
        'Emission Diff (%)': 'Emission Difference'
    }

    # Define the desired column order
    column_order = ['Departure Date', 'Arrival Date', 'From', 'To', 'Airline',
                    'Duration', 'Price', 'Number of Stops', 'Layover Time',
                    'Accessed Date', 'CO2 Emission', 'Emission Difference']

    # Rename columns and reorder them
    flight_data_renamed = flight_data.rename(columns=new_column_names)[column_order]

    # Output
    print("-------Output-----------------------------------------------------------------------------")
    print(flight_data_renamed)
    print("------------------------------------------------------------------------------------------")

def find_lowest_price_rows(dataframe):
    # Check if the dataframe is empty
    if dataframe.empty:
        return None, None

    # Split the DataFrame into up and down flights
    up_flights = dataframe[dataframe['Direction'] == 'Up']
    down_flights = dataframe[dataframe['Direction'] == 'Down']

    # Find the row with the lowest price for up and down flights
    lowest_price_up = up_flights.loc[up_flights['Price ($)'].idxmin()]
    lowest_price_down = down_flights.loc[down_flights['Price ($)'].idxmin()]

    return lowest_price_up, lowest_price_down
def save_flight_data_to_excel(flight_data, filename):
    # Define the new column names and order
    new_column_names = {
        'Departure datetime': 'Departure Date',
        'Arrival datetime': 'Arrival Date',
        'Origin': 'From',
        'Destination': 'To',
        'Airline(s)': 'Airline',
        'Travel Time': 'Duration',
        'Price ($)': 'Price',
        'Num Stops': 'Number of Stops',
        'Layover': 'Layover Time',
        'Access Date': 'Accessed Date',
        'CO2 Emission (kg)': 'CO2 Emission',
        'Emission Diff (%)': 'Emission Difference'
    }

    # Define the desired column order
    column_order = ['Departure Date', 'Arrival Date', 'From', 'To', 'Airline',
                    'Duration', 'Price', 'Number of Stops', 'Layover Time',
                    'Accessed Date', 'CO2 Emission', 'Emission Difference']

    # Rename columns and reorder them
    flight_data_renamed = flight_data.rename(columns=new_column_names)[column_order]

    # Save DataFrame to Excel
    flight_data_renamed.to_excel(filename, index=False)
def add_direction_column(dataframe):
    # Check if the DataFrame is empty
    if dataframe.empty:
        return dataframe

    # Initialize an empty list to store the directions
    directions = []

    # Iterate over the DataFrame rows
    for index, row in dataframe.iterrows():
        # Check if the destination airport matches the first origin airport
        if row['Destination'] == dataframe.iloc[0]['Origin']:
            directions.append('Down')
        else:
            directions.append('Up')

    # Add the directions list as a new column to the DataFrame
    dataframe['Direction'] = directions

    return dataframe
def scrape_flight_data(origin, destination, start_date, end_date):
    # Obtain our scrape object, represents our query
    result = Scrape(origin, destination, start_date, end_date)

    # Run selenium through ChromeDriver, modifies results in-place
    ScrapeObjects(result)

    # Return the queried representation of result
    return result.data

def generate_date_combinations(start_date,end_date,start_interval=0, end_interval=0):
    date_combinations = []
    current_start_date = start_date
    current_end_date = end_date

    while current_start_date < current_end_date:
        current_start = current_start_date.strftime('%Y-%m-%d')
        current_end = current_end_date.strftime('%Y-%m-%d')
        date_combinations.append((current_start, current_end))
        current_start_date += timedelta(days=start_interval)
        current_end_date += timedelta(days=end_interval)

    return date_combinations # [('2024-07-20', '2024-08-20'), ('2024-07-22', '2024-08-23'), ('2024-07-24', '2024-08-26'),.......]


def main():
    # Input parameters
    destination = 'BBI'  # Origin
    origin = 'DFW'  # Destination
    start_date_main_str = '2024-05-02'
    end_date_main_str = '2024-05-20'

    # Convert start_date_main_str and end_date_main_str to datetime objects
    start_date_main = datetime.strptime(start_date_main_str, '%Y-%m-%d')
    end_date_main = datetime.strptime(end_date_main_str, '%Y-%m-%d')

    # Define start and end intervals
    start_interval = 1
    end_interval = 0

    # Generate date combinations
    date_combinations = generate_date_combinations(start_date_main, end_date_main, start_interval, end_interval)

    print("Date Combinations:")
    for start_date, end_date in date_combinations:
        print(f"Start Date: {start_date}, End Date: {end_date}")

    f_name = generate_filename(start_date_main, end_date_main, start_interval, end_interval)
    print("Generated filename:", f_name)

    # Iterate over date combinations
    for start_date, end_date in date_combinations:
        print("\n----------------------------------------")
        print(f"Scraping flight data for period {start_date} to {end_date}...")
        # Scrape flight data
        flight_data = scrape_flight_data(origin, destination, start_date, end_date)
        # Add Direction Column
        updated_flight_data = add_direction_column(flight_data)
        # Find lowest price for up and down flights
        up_lowest_price, down_lowest_price = find_lowest_price_rows(flight_data)

        # Print lowest price details
        print("Lowest Price Details:")
        print(f"Up flight: {up_lowest_price}")
        print(f"Down flight: {down_lowest_price}")

        # Concatenate the 'Up' and 'Down' DataFrames
        up_lowest_price_table = FlightData('up', up_lowest_price)
        down_lowest_price_table = FlightData('down', down_lowest_price)

        # Write to Excel
        print(f"Writing to Excel file: {f_name}")
        write_to_excel(up_lowest_price_table, down_lowest_price_table, f_name)
        print("Data Saved to Excel.")
        print("----------------------------------------\n")


if __name__ == "__main__":
    main()

if __name__ == "__main__":
    main()
